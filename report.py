import requests
import urllib.request, urllib.parse
import configparser
import openpyxl
import string
import logging
from abc import ABC
from pathlib import Path
from datetime import datetime
import shutil
from itertools import groupby
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font, Alignment

PYTHONDONTWRITEBYTECODE=1

logging.basicConfig(
    filename="log/app.log",
    filemode="a",
    format="%(asctime)s,%(msecs)d %(name)s %(levelname)s %(message)s",
    datefmt="%H:%M:%S",
    encoding="utf-8",
    level=logging.DEBUG,
)

logger = logging.getLogger(__name__)


# № заявки n_abs
# дата поступления timez
# наименование объекта nameobj
# номер объекта numobj
# адрес объекта addrobj
# техник tehn
# назначенное время timev


class Config(ABC):
    """Базовый класс для реализации методов получения параметров"""

    default_encoding = "utf-8"
    """ Кодировка файла конфигурации """

    def __init__(self, config_file_path: str = "config.ini") -> None:
        self.config = configparser.ConfigParser()
        self.config.read(config_file_path, encoding=self.default_encoding)


class CobraConfig(Config):
    """Получение параметров, отвечающих за подключение к КПО Кобра"""

    section = "Cobra"
    """ Название секции файла конфигурации """

    host_param = "host"
    """ Имя параметра, хранящего адрес или FQDN-имя хоста для подключения к КПО Кобра """

    port_param = "port"
    """ Имя параметра, хранящего порт для подключения к КПО КОбра """

    token_param = "token"
    """ Имя параметра, хранящего пароль удаленного доступа pud для подключения к КПО Кобра """

    def get_host(self) -> str:
        """Возвращает адрес хоста или FQDN-имя сервера КПО Кобра"""
        return self.config.get(self.section, self.host_param)

    def get_port(self) -> str:
        """Возвращает порт для подключения к КПО Кобра"""
        return self.config.get(self.section, self.port_param)

    def get_token(self) -> str:
        """Возвращает пароль удаленного доступа pud для подключения к КПО Кобра"""
        return self.config.get(self.section, self.token_param)


class TelegramConfig(Config):
    """Получение параметров, отвечающих за взаимодействие с Telegram API"""

    section = "Telegram"
    """ Название секции файла конфигурации """

    token_param = "token"
    """ Имя параметра, хранящего данные токена бота """

    chat_id_param = "chat_id"
    """ Имя параметра, хранящего данные с ID чата в Telegram """

    def get_token(self) -> str:
        """Возвращает токен бота Telegram"""
        return self.config.get(self.section, self.token_param)

    def get_chat_id(self) -> str:
        """Возвращает ID чата в Telegram для отправки уведомлений"""
        return self.config.get(self.section, self.chat_id_param)


class CobraTaskReportHeader:
    """Объект передачи данных, содержащий соответствие названий заголовков
    таблицы полям, возвращаемым REST API КПО Кобра"""

    n_abs = "№ заявки"
    timez = "Дата поступления"
    # account_number = "Пультовый номер"
    nameobj = "Наименование объекта"
    numobj = "Номер объекта"
    addrobj = "Адрес объекта"
    tehn = "Техник"
    timev = "Назначенное время"


class ExcelReport(ABC):
    """Базовый класс для генерации отчетов КПО Кобра."""

    export_path = "out/"
    """ Каталог экспорта файлов относительно каталога текущего проекта """

    file_suffix = "xlsx"
    """ Расширение экспортируемого файла """

    def __init__(self) -> None:
        self._workbook = openpyxl.Workbook()
        self._sheet = self._workbook.active
        self._columns = string.ascii_uppercase
        self.file_name = ""

    def save(self):
        """Сохраняет экспортирвуемый файл"""
        self.export_filename = self._set_file_name()
        self._workbook.save(self.export_filename)

    def _set_file_name(self) -> Path:
        """Генерация пути к файлу экспорта
        Имя файла экспорта должно задаваться атрибутом file_name в классе-потомке
        """
        self._create_export_path_if_not_exists()
        current_datetime = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
        return Path(
            self.export_path
            + current_datetime
            + "_"
            + self.file_name
            + "."
            + self.file_suffix
        )

    def _create_export_path_if_not_exists(self):
        """Создает каталог экспорта в случае его отсутствия"""
        Path(self.export_path).mkdir(parents=True, exist_ok=True)


class CobraTaskExcelReport(ExcelReport):
    """Реализация отчета, содержащего данные открытых заявок"""

    file_name = "Оперативные_заявки"
    """ Имя генерируемого файла без расширения """

    max_row = 1
    """ Номер 1-й строки отчета (итерируется) """

    def set_row(self, task: dict):
        """Запись данных одной заявки в таблицу

        Args:
            task (dict): словарь, содержащий данные одной заявки, полученной в
            ответе от REST API КПО Кобра
        """
        for key, value in task.items():
            column_index = list(task.keys()).index(key)
            """ Для 1й записи добавляется заголовок """
            thin_border = Side(border_style="thin", color="000000")
            border = Border(
                top=thin_border, left=thin_border, right=thin_border, bottom=thin_border
            )
            if self.max_row == 1:
                table_title = getattr(CobraTaskReportHeader, key)
                self._sheet[self._columns[column_index] + str(self.max_row)] = (
                    table_title
                )
                self._sheet[self._columns[column_index] + str(self.max_row)].border = (
                    border
                )
                self._sheet[self._columns[column_index] + str(self.max_row)].font = (
                    Font(bold=True)
                )

                self._sheet[self._columns[column_index] + str(self.max_row + 1)] = str(
                    value
                )
                self._sheet[
                    self._columns[column_index] + str(self.max_row + 1)
                ].border = border
                self._sheet[
                    self._columns[column_index] + str(self.max_row + 1)
                ].alignment = Alignment(wrap_text=True)
            else:
                self._sheet[self._columns[column_index] + str(self.max_row)] = str(
                    value
                )
                self._sheet[self._columns[column_index] + str(self.max_row)].border = (
                    border
                )
                self._sheet[
                    self._columns[column_index] + str(self.max_row)
                ].alignment = Alignment(wrap_text=True)
        """ Установка ширины столбцов """
        self._set_columns_width()
        """ Увеличение счетчика номера строки для следующей итерации. 
        Увеличение на 2 для 1-й записи, т.к. добавляются заголовки таблицы """
        self.max_row += 1 if self.max_row > 1 else 2

    def _set_columns_width(self):
        self._sheet.column_dimensions["A"].width = 15
        self._sheet.column_dimensions["B"].width = 20
        self._sheet.column_dimensions["C"].width = 20
        self._sheet.column_dimensions["D"].width = 25
        self._sheet.column_dimensions["E"].width = 20
        self._sheet.column_dimensions["F"].width = 25
        self._sheet.column_dimensions["G"].width = 30


class CobraTable(ABC):
    """Базовый класс для получения данных из таблиц КПО Кобра"""

    endpoint_root = "api.table.get"
    """ Метод для получения данных таблиц """

    token_key = "pud"
    """ Наименование параметра, в котором передается пароль удаленного доступа """

    def __init__(self, config: CobraConfig) -> None:
        self._host = config.get_host()
        self._port = config.get_port()
        self._token = config.get_token()
        self._endpoint_url = self._get_endpoint_url()

    def _get_endpoint_url(self) -> str:
        """Генерирует полный url для отправки http-запроса"""
        endpoint = f"{self._host}:{self._port}/{self.endpoint_root}"
        token = urllib.parse.urlencode({self.token_key: self._token})
        return f"{endpoint}?{token}"


class CobraTaskReport(CobraTable):
    """Реализует запрос к REST API КПО Кобра, формирует параметры запроса
    (фильтр, набор возвращаемых полей)

    TODO генерировать набор возвращаемых полей из объекта передачи данных,
    хранящего названия заголовков таблицы отчета
    """

    name_template = "***"
    """ Шаблон наименования заявки. В отчет попадают только заявки, формируемые
    оперативным дежурным начинаются с *** """

    table_name = "zayavki"
    """ Название таблицы, хранящей данные заявок """

    def get_tasks(self) -> tuple:
        """Получает заявки из КПО Кобра"""
        response = self._get_unfinished_tasks()
        return tuple(response)

    def _get_unfinished_tasks(self):
        """Запрос текущих заявок из КПО Кобра"""
        url = f"{self._endpoint_url}"
        params = {
            "name": self.table_name,
            "filter": self._get_filter(),
            "fields": self._get_fields(),
        }
        response = requests.get(url=url, params=params).json()
        result = sorted(response["result"], key=lambda task: task["tehn"])
        return result

    def _get_filter(self):
        """Установка фильтра при запросе заявок из КПО Кобра"""
        filter_value = '[{"zay": "' + self.name_template + '"}]'
        return f"{filter_value}"

    def _get_fields(self):
        """Запрос полей из таблицы КПО Кобра, соответствующих
        формату генерируемого отчета"""
        fields_value = '[{"n_abs": "1"}, {"timez": "1"}, {"nameobj": "1"}, {"numobj": "1"}, {"addrobj": "1"}, {"tehn": "1"}, {"timev": "1"}]'
        return f"{fields_value}"


class CobraTaskReportMessage:
    """Генерация текста сообщения отчета на основании данных
    заявки из КПО Кобра"""

    message: str = str()
    """ Пустая строка сообщения """

    def add_report_header(self) -> None:
        """Добавляет заголовок к сообщению отчета"""
        current_date = datetime.today().strftime("%d.%m.%Y")
        self.message += f"Распоряжение оперативного дежурного на {current_date}"
        self.add_empty_string_to_report_message()
        self.message += "Оперативные заявки"
        self.add_empty_string_to_report_message()
        self.add_empty_string_to_report_message()

    def add_tehn_to_report_message(self, tehn: str) -> None:
        """Добавление имени техника в сообщение отчета

        Args:
            tehn (str): Ф.И.О. техника, закрепленного за заявкой из КПО Кобра
        """
        self.message += str(tehn)
        self.add_empty_string_to_report_message()
        self.add_empty_string_to_report_message()

    def add_task_to_report_message(self, task: dict) -> None:
        """Добавление данных одной заявки, полученной из КПО Кобра,
        в строку сообщения отчета

        Args:
            task (dict): словарь содержащий данные одной заявки, полученный из
            КПО Кобра
        """
        task_string = f"{task['numobj']} {task['nameobj']} {task['addrobj']}"
        self.message += str(task_string)
        self.add_empty_string_to_report_message()

    def add_empty_string_to_report_message(self) -> None:
        """Добавляет пустую строку в текст сообщения отчета"""
        self.message += "\r\n"

    def get_report_message_text(self) -> str:
        """Возвращает сгенерированную строку сообщения отчета

        Returns:
            str: строка сообщения отчета
        """
        return self.message


class Telegram:
    """Отправка информации в Telegram"""

    base_url = "https://api.telegram.org/bot"
    """ Базовый URL Telegram Bot API для направления запросов """

    text_message_endpoint = "/sendMessage"
    """ Метод REST API Telegram Bot API для отправки текстовых сообщений """

    document_attachment_endpoint = "/sendDocument"
    """ Метод REST API Telegram Bot API для отправки файлов """

    def __init__(self, config: TelegramConfig) -> None:
        self._token = config.get_token()
        self._chat_id = config.get_chat_id()

    def send_text_message(self, text: str) -> None:
        """Отправка текстового сообщения

        Args:
            text (str): текст отправляемого сообщения
        """
        url = f"{self.base_url}{self._token}{self.text_message_endpoint}"
        params = {"chat_id": self._chat_id, "text": text}
        logger.info("Отправка текстового сообщения в Telegram")
        requests.get(url=url, params=params).json()

    def send_document_attachment(self, document_path: str) -> None:
        """Отправка документа

        Args:
            document_path (str): путь к отправляемому документу
        """
        files = {"document": open(document_path, "rb")}
        url = f"{self.base_url}{self._token}{self.document_attachment_endpoint}"
        data = {
            "chat_id": self._chat_id,
        }
        logger.info("Отправка отчета в Telegram")
        requests.post(url=url, data=data, files=files)


def main():
    # Запрос заявок из КПО Кобра
    cobra_config = CobraConfig()
    cobra_base = CobraTaskReport(cobra_config)
    task_objects = cobra_base.get_tasks()
    if len(task_objects):
        # Генерация файла отчета и текста сообщения в Telegram
        task_report = CobraTaskExcelReport()
        report_message = CobraTaskReportMessage()
        report_message.add_report_header()
        for tehn, one_tehn_tasks in groupby(
            task_objects, lambda task_list: task_list["tehn"]
        ):
            report_message.add_tehn_to_report_message(tehn)
            for task in one_tehn_tasks:
                report_message.add_task_to_report_message(task)
                task_report.set_row(task)
            report_message.add_empty_string_to_report_message()
        task_report.save()

        # Отправка уведомлений Telegram
        tg_config = TelegramConfig()
        tg = Telegram(tg_config)
        tg.send_text_message(report_message.get_report_message_text())
        tg.send_document_attachment(task_report.export_filename)
    else:
        logger.info("Заявки отсутствуют")

    # Очистка каталог экспорта отчетов
    shutil.rmtree(
        ExcelReport.export_path,
        ignore_errors=True,
        onerror=logger.warning("Попытка очистки каталога завершилась ошибкой"),
    )


if __name__ == "__main__":
    main()
