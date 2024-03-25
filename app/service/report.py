"""Генерация отчетов.

Логика генерации отчетов. В формате текстовых сообщений Telegram и Excel.
"""

# Standard Library
import string
from abc import ABC
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.styles.borders import Border, Side

from app.service.cobra import TaskReportHeader


class ExcelReport(ABC):
    """Базовый класс для генерации отчетов КПО Кобра."""

    export_path = "job/"
    """ Каталог экспорта файлов относительно каталога текущего проекта """

    file_suffix = "xlsx"
    """ Расширение экспортируемого файла """

    def __init__(self) -> None:  # noqa D107
        self._workbook = openpyxl.Workbook()
        self._sheet = self._workbook.active
        self._columns = string.ascii_uppercase

    def save(self):
        """Сохраняет экспортирвуемый файл."""
        self.export_filename = self._set_file_name()
        self._workbook.save(self.export_filename)

    def _set_file_name(self) -> Path:
        """Генерация пути к файлу экспорта.

        Имя файла экспорта задается атрибутом file_name в классе-потомке.
        """
        # self._create_export_path_if_not_exists()
        current_datetime = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
        file_path = Path(
            f"{self.export_path}{current_datetime}_\
{self.file_name}.{self.file_suffix}"  # type: ignore
        )
        return file_path

        # def _create_export_path_if_not_exists(self):
        """Создает каталог экспорта в случае его отсутствия."""
        # Path(self.export_path).mkdir(parents=True, exist_ok=True)


class CobraTaskExcelReport(ExcelReport):
    """Реализация отчета, содержащего данные открытых заявок."""

    file_name = "Аварийные_заявки"
    """ Имя генерируемого файла без расширения """

    max_row = 3
    """ Номер 1-й строки отчета (итерируется) """

    start_row = 3

    def set_header(self):
        """Устанавливает заголовок таблицы отчета в формате Excel."""
        current_date = datetime.today().strftime("%d.%m.%Y")
        self._sheet.merge_cells("A1:H1")
        self._sheet["A1"] = f"Аварийные Заявки {current_date}"
        bold_font = Font(bold=True)
        self._sheet["A1"].font = bold_font
        self._sheet["A1"].alignment = Alignment(horizontal="center")

    def set_footer(self):
        """Устанавливает дисклеймер отчета в формате Excel."""
        current_datetime = datetime.now().strftime("%d.%m.%Y %H:%M:%S")
        self._sheet.merge_cells(
            "A" + str(self.max_row + 1) + ":H" + str(self.max_row + 1)
        )
        self._sheet["A" + str(self.max_row + 1)] = (
            f"Дата формирования отчета: {current_datetime}"
        )

    def set_row(self, task: dict) -> None:
        """Устанавливает данные строки отчета."""
        if self.max_row == self.start_row:
            self._sheet["A" + str(self.max_row)] = TaskReportHeader().timez
            self._sheet["B" + str(self.max_row)] = TaskReportHeader().timev
            self._sheet["C" + str(self.max_row)] = TaskReportHeader().prin
            self._sheet["D" + str(self.max_row)] = TaskReportHeader().numobj
            self._sheet["E" + str(self.max_row)] = TaskReportHeader().nameobj
            self._sheet["F" + str(self.max_row)] = TaskReportHeader().addrobj
            self._sheet["G" + str(self.max_row)] = TaskReportHeader().zay
            self._sheet["H" + str(self.max_row)] = TaskReportHeader().tehn
            self._set_row_border(self.max_row)
            self._set_font_bold(self.max_row)
            self._add_row_by_number(task, self.max_row + 1)
            self._set_row_border(self.max_row + 1)
            self._set_text_alignment(self.max_row + 1, True)
        else:
            self._add_row_by_number(task, self.max_row)
            self._set_row_border(self.max_row)
            self._set_text_alignment(self.max_row, True)

        """ Установка ширины столбцов """
        self._set_columns_width()
        """ Увеличение счетчика номера строки для следующей итерации.
        Увеличение на 2 для 1-й записи, т.к. добавляются заголовки таблицы """
        self.max_row += 1 if self.max_row > self.start_row else 2

    def _add_row_by_number(self, task: dict, number: int):
        self._sheet["A" + str(number)] = task["timez"]
        self._sheet["B" + str(number)] = task["timev"]
        self._sheet["C" + str(number)] = task["prin"]
        self._sheet["D" + str(number)] = task["numobj"]
        self._sheet["E" + str(number)] = task["nameobj"]
        self._sheet["F" + str(number)] = task["addrobj"]
        self._sheet["G" + str(number)] = task["zay"]
        self._sheet["H" + str(number)] = task["tehn"]

    def _set_row_border(self, number: int) -> None:
        brdr = Side(border_style="thin", color="000000")
        border = Border(top=brdr, left=brdr, right=brdr, bottom=brdr)
        self._sheet["A" + str(number)].border = border
        self._sheet["B" + str(number)].border = border
        self._sheet["C" + str(number)].border = border
        self._sheet["D" + str(number)].border = border
        self._sheet["E" + str(number)].border = border
        self._sheet["F" + str(number)].border = border
        self._sheet["G" + str(number)].border = border
        self._sheet["H" + str(number)].border = border

    def _set_text_alignment(self, number: int, wrap: bool = False):
        align = Alignment(wrap_text=wrap, horizontal="general", vertical="top")
        self._sheet["A" + str(number)].alignment = align
        self._sheet["B" + str(number)].alignment = align
        self._sheet["C" + str(number)].alignment = align
        self._sheet["D" + str(number)].alignment = align
        self._sheet["E" + str(number)].alignment = align
        self._sheet["F" + str(number)].alignment = align
        self._sheet["G" + str(number)].alignment = align
        self._sheet["H" + str(number)].alignment = align

    def _set_font_bold(self, number: int):
        bold_font = Font(bold=True)
        self._sheet["A" + str(number)].font = bold_font
        self._sheet["B" + str(number)].font = bold_font
        self._sheet["C" + str(number)].font = bold_font
        self._sheet["D" + str(number)].font = bold_font
        self._sheet["E" + str(number)].font = bold_font
        self._sheet["F" + str(number)].font = bold_font
        self._sheet["G" + str(number)].font = bold_font
        self._sheet["H" + str(number)].font = bold_font

    def _set_columns_width(self):
        self._sheet.column_dimensions["A"].width = 20
        self._sheet.column_dimensions["B"].width = 20
        self._sheet.column_dimensions["C"].width = 25
        self._sheet.column_dimensions["D"].width = 20
        self._sheet.column_dimensions["E"].width = 25
        self._sheet.column_dimensions["F"].width = 30
        self._sheet.column_dimensions["G"].width = 30
        self._sheet.column_dimensions["H"].width = 30
